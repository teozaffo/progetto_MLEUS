# app.py
from openpyxl import load_workbook
import pytz
from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
import pandas as pd
import os
import json
import uuid
import jwt
from datetime import datetime
from model_prediction import predict_input
import pytz
from openpyxl import load_workbook, Workbook
from dateutil import parser
from dotenv import load_dotenv


app = Flask(__name__)
CORS(app)

load_dotenv()

excel_file_path = os.getenv('EXCEL_FILE')
jwt_secret = os.getenv('JWT_SECRET')

'''
columns = [
    "Datetime",
    "Ip Address",
    "Model",
    "Hospital Center",
    "Protocol Code",
    "Age",
    "Sex",
    "Max Dim",
    "Min Dim",
    "Veinous Inf",
    "Arterious Inf",
    "Duct Ret/Ductal Inv",
    "Vessel Comp",
    "Lymphadenopathy",
    "Reg Margins",
    "Echogenicity",
    "Mult Lesions",
    "Prediction"
]

if os.path.exists(excel_file_path_path):
    excel_df = pd.read_excel(excel_file_path_path)
    print("📂 File Excel trovato con", len(df), "righe")
else:
    excel_df = pd.DataFrame(columns=columns)
    print("📁 File non trovato, ne creo uno nuovo con intestazioni")
'''


@app.route("/")
def index():
    return render_template("index.html")

@app.route("/")
def home():
    return "✅ Server Flask attivo su PythonAnywhere!"
    

@app.route("/prediction")
def prediction():
    print("now at prediction")
    return render_template("prediction.html")


@app.route("/prediction")
def show_prediction():
    print("now at prediction")
    return render_template("prediction.html")


# Route per salvare i dati nel file Excel
@app.route("/salva", methods=["POST"])
def salva_diagnosi():
    data = request.get_json()
    print("✅ Richiesta ricevuta:", data)

    # Costruisci la nuova riga da salvare
    new_row = {
        "ID": str(uuid.uuid4()),
        "Datetime": datetime.now(pytz.timezone("Europe/Rome")).strftime("%Y-%m-%d %H:%M:%S"),
        "Ip Address": data.get("ip", request.remote_addr),
        "Model": data.get("model", "unknown"),
        "Hospital Center": data.get("HospitalCenter", ""),
        "Protocol Code": data.get("ProtocolCode", ""),
        "Age": data.get("age"),
        "Sex": data.get("sex"),
        "Max Dim": data.get("Dim1"),
        "Min Dim": data.get("Dim2"),
        "Veinous Inf": data.get("Veins"),
        "Arterious Inf": data.get("Arteries"),
        "Duct Ret/Ductal Inv": data.get("DuctRetrodilatation"),
        "Vessel Comp": data.get("VesselCompression"),
        "Lymphadenopathy": data.get("Lymphadenopathy"),
        "Reg Margins": data.get("Margins"),
        "Echogenicity": data.get("Ecostructure"),
        "Mult Lesions": data.get("Multiple"),
        "Prediction": data.get("prediction") 
    }

    print("📄 Riga da salvare:", new_row)

    try:
        print(f"ID: {new_row['ID']}")
        
        # Se il file non esiste, crea un nuovo DataFrame con intestazioni
        if os.path.exists(excel_file_path):
            df = pd.read_excel(excel_file_path)
            print("📂 File Excel trovato con", len(df), "righe")
        else:
            df = pd.DataFrame(columns=new_row.keys())
            print("📁 File non trovato, ne creo uno nuovo con intestazioni")

        # Aggiungi la nuova riga
        # df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        new_entry = pd.DataFrame([new_row])
        if not df.empty:
            new_entry = pd.DataFrame([new_row], columns=df.columns)
        df = pd.concat([df, new_entry], ignore_index=True)

        # Salva il file
        df.to_excel(excel_file_path, index=False)
        # Allarga automaticamente le colonne
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_file_path)
        ws = wb.active

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = length + 2  # margine extra
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(excel_file_path)
        
        token = jwt.encode({'ID': new_row["ID"]}, jwt_secret, algorithm='HS256')

        print("💾 Dati scritti correttamente su", excel_file_path)

        return jsonify({
            "message": "Dati salvati con successo",
            "token": token
        })
    except Exception as e:
        print("❌ Errore nella scrittura su Excel:", str(e))
        return jsonify({"error": str(e)}), 500

    
@app.route("/questionario", methods=["POST"])
def ricevi_questionario():
    data = request.get_json()
    print("📩 Ricevuto questionario:", data)

    try:
        df = pd.read_excel(excel_file_path)

        # Confronto robusto con datetime localizzato
        incoming_dt = parser.isoparse(data["datetime"]).astimezone(pytz.timezone("Europe/Rome"))
        incoming_str = incoming_dt.strftime("%Y-%m-%d %H:%M")
        
        prefix = 'Bearer '
        
        incoming_token = request.headers['Authorization'][len(prefix):]
        
        request_id = jwt.decode(incoming_token, jwt_secret, algorithms='HS256')["ID"]

        print(f"ID: {request_id}")

        # Conversione della colonna Datetime
        df["Datetime"] = pd.to_datetime(df["Datetime"])
        excel_times = df["Datetime"].dt.strftime("%Y-%m-%d %H:%M")
        idx = excel_times[excel_times == incoming_str].index
        
        excel_ids = df["ID"]
        
        id_idx = excel_ids[excel_ids == request_id].index

        if len(idx) == 0:
            print("❌ Riga non trovata per datetime:", data["datetime"])
            return jsonify({"error": "Riga non trovata"}), 404

        i = id_idx[0]

        df.at[i, "Q1"] = int(data.get("q1")) if data.get("q1") is not None else None
        df.at[i, "Q2"] = int(data.get("q2")) if data.get("q2") is not None else None
        df.at[i, "Q3"] = int(data.get("q3")) if data.get("q3") is not None else None
        df.at[i, "Q4"] = int(data.get("q4")) if data.get("q4") is not None else None
        df.at[i, "Q5"] = str(data.get("q5")) if data.get("q5") is not None else ""


        df.to_excel(excel_file_path, index=False)
        print("✅ Questionario salvato con successo")
        
        # Ridimensiona le colonne dopo aver salvato i nuovi dati
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_file_path)
        ws = wb.active

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = length + 2
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(excel_file_path)
        return jsonify({"message": "Salvato con successo"})

    except Exception as e:
        print("❌ Errore durante il salvataggio del questionario:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route("/model_predict", methods=["POST"])
def model_predict():
    data = request.get_json()
    try :
        result, color = predict_input(data)
        
        percent = round(float(result))       
        data['prediction'] = f"{percent}% Malignant"
        save_response = salva_diagnosi().get_json()
        
        print(save_response)
        
        return jsonify({
            "message": f"case Predicted and Saved Succesfully!",
            "token": save_response['token'],
            "prediction": str(percent),
            "backgroundColor": color
        }) 
    except Exception as e:
        print(f"Encountered Error: {str(e)}")
        return jsonify({"error": str(e)}), 500
