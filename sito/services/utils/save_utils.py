from datetime import datetime
import pytz
import os
from dateutil import parser
from flask import request
import pandas as pd
import jwt
from openpyxl import load_workbook

excel_file_path = "./diagnosi.xlsx"




def parse_new_row(data):
  print(data)
  return {
    "Datetime": datetime.now(pytz.timezone("Europe/Rome")).strftime("%Y-%m-%d %H:%M:%S"),
    "Ip Address": data.get("Ip"),
    "Model": data.get("model", "unknown"),
    "Hospital Center": data.get("HospitalCenter", ""),
    "Protocol Code": data.get("ProtocolCode", ""),
    "Age": data.get("age"),
    "Sex": data.get("sex"),
    "Max Dim": data.get("Dim1"),
    "Min Dim": data.get("Dim2"),
    "Veinous Inf": data.get("Veins"),
    "Arterious Inf": data.get("Arteries"),
    "Duct Ret/Ductal Inv": data.get("DuctRetrodilation"),
    "Vessel Comp": data.get("VesselCompression"),
    "Lymphadenopathy": data.get("Lymphadenopathy"),
    "Reg Margins": data.get("Margins"),
    "Echogenicity": data.get("Ecostructure"),
    "Mult Lesions": data.get("Multiple"),
    "Prediction": data.get("prediction")
  }
  
  



def add_new_row_to_excel(new_row):
  # Se il file non esiste, crea un nuovo DataFrame con intestazioni
  if os.path.exists(excel_file_path):
      df = pd.read_excel(excel_file_path)
      print("📂 File Excel trovato con", len(df), "righe")
  else:
      df = pd.DataFrame(columns=new_row.keys())
      print("📁 File non trovato, ne creo uno nuovo con intestazioni")

  new_entry = pd.DataFrame([new_row])
  if not df.empty:
      new_entry = pd.DataFrame([new_row], columns=df.columns)
  df = pd.concat([df, new_entry], ignore_index=True)

  # Salva il file
  df.to_excel(excel_file_path, index=False)
  # Allarga automaticamente le colonne
  from openpyxl.utils import get_column_letter

  wb = load_workbook(excel_file_path)
  ws = wb.active

  for column_cells in ws.columns:
      length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
      adjusted_width = length + 2  # margine extra
      col_letter = get_column_letter(column_cells[0].column)
      ws.column_dimensions[col_letter].width = adjusted_width

  wb.save(excel_file_path)

  print("💾 Dati scritti correttamente su", excel_file_path)
  
  
  
  
      
def add_feddback_to_existing_row(data, user_token):
  df = pd.read_excel(excel_file_path)
  
  date_time = get_datetime_from_user_token(user_token=user_token)

  # Confronto robusto con datetime localizzato
  incoming_dt = parser.isoparse(date_time["Datetime"])
  incoming_str = incoming_dt.strftime("%Y-%m-%d %H:%M")


  # Conversione della colonna Datetime
  df["Datetime"] = pd.to_datetime(df["Datetime"])
  excel_times = df["Datetime"].dt.strftime("%Y-%m-%d %H:%M")
  print("Incoming datetime string:", incoming_str)
  print("Excel datetime strings (tail):")
  print(excel_times.tail(10).to_list())
  idx = excel_times[excel_times == incoming_str].index

  if len(idx) == 0:
      print("❌ Riga non trovata per datetime:", incoming_str)
      raise Exception("Riga non trovata")

  i = idx[0]

  df.at[i, "Q1"] = int(data.get("q1")) if data.get("q1") is not None else None
  df.at[i, "Q2"] = int(data.get("q2")) if data.get("q2") is not None else None
  df.at[i, "Q3"] = int(data.get("q3")) if data.get("q3") is not None else None
  df.at[i, "Q4"] = int(data.get("q4")) if data.get("q4") is not None else None
  df.at[i, "Q5"] = str(data.get("q5")) if data.get("q5") is not None else ""


  df.to_excel(excel_file_path, index=False)
  print("✅ Questionario salvato con successo")
  
  # Ridimensiona le colonne dopo aver salvato i nuovi dati
  from openpyxl.utils import get_column_letter

  wb = load_workbook(excel_file_path)
  ws = wb.active

  for column_cells in ws.columns:
      length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
      adjusted_width = length + 2
      col_letter = get_column_letter(column_cells[0].column)
      ws.column_dimensions[col_letter].width = adjusted_width

      wb.save(excel_file_path)
      
      
def generate_user_token_from_datetime(date_time):
  jwt_secret = os.getenv("JWT_SECRET")
  jwt_algo = os.getenv("JWT_ALGO")
  
  user_token = jwt.encode({"Datetime": date_time}, jwt_secret, jwt_algo)
  
  return user_token
  


def get_datetime_from_user_token(user_token):
  jwt_secret = os.getenv("JWT_SECRET")
  jwt_algo = os.getenv("JWT_ALGO")
  
  return jwt.decode(user_token, jwt_secret, jwt_algo)
